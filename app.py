import io
import re
import unicodedata
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


st.set_page_config(page_title="NHO Attendance Checker", page_icon="✅", layout="wide")

FIELD_ALIASES = {
    "Applicant Name": ["applicant name", "talent name"],
    "Talent Email": ["talent email", "email"],
    "Current Onboarding Stage": ["current onboarding stage"],
    "HP Overview Template": ["hp overview template"],
    "OB365": ["ob365"],
    "District Release": ["district release"],
    "Pre-hire Training": [
        "pre hire training", "prehire training",
        "classmarts", "classmarts cert",
        "classsmarts", "classsmarts cert",
    ],
    "Safety & Security": ["safety security", "policy explainer cert"],
    "Onboarding Introduction Email": ["onboarding introduction email"],
    "I-9": ["i 9", "i9"],
    "Education Requirement": ["education requirement"],
    "Certified": ["certified", "yes"],
    "Drug Screen": ["drug screen", "drug screem"],
    "Background Check": ["background check"],
    "Fingerprint Screening": ["fp screening", "scheduled"],
    "MDPS Training": ["mdps training"],
    "Badge Photo": ["badge photo"],
    "Notes": ["notes", "notes all calls email must be logged in ksn bh"],
}

REQUIREMENTS = [
    "I-9",
    "Drug Screen",
    "Fingerprint Screening",
]

ARCHIVE_SHEETS = {
    "landing page", "drop downs", "template", "call back", "transitions",
    "inactivated during hiring", "newly activated", "rejected",
}

EMAIL_PATTERN = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def clean_header(value):
    text = normalize_text(value)
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def normalize_text(value):
    if value is None or pd.isna(value):
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().lower()


def normalize_name(value):
    return re.sub(r"[^a-z0-9 ]", "", normalize_text(value))


def normalize_email(value):
    return re.sub(r"\s+", "", normalize_text(value))


def find_header_row(rows):
    for index, row in enumerate(rows[:12]):
        headers = {clean_header(value) for value in row if value is not None}
        if "applicant name" in headers and ("talent email" in headers or "ksn id" in headers):
            return index
    return None


def canonical_column(header):
    candidate = clean_header(header)
    for canonical, aliases in FIELD_ALIASES.items():
        if any(candidate == alias or candidate.startswith(alias + " ") for alias in aliases):
            return canonical
    return None


def infer_email_column(rows, header_index):
    """Find the most email-like column when the tracker header is misspelled."""
    scores = {}
    for row in rows[header_index + 1:header_index + 51]:
        for index, value in enumerate(row):
            if EMAIL_PATTERN.match(normalize_email(value)):
                scores[index] = scores.get(index, 0) + 1
    return max(scores, key=scores.get) if scores else None


@st.cache_data(show_spinner=False)
def read_tracker(file_bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    records = []
    for worksheet in workbook.worksheets:
        if normalize_text(worksheet.title) in ARCHIVE_SHEETS:
            continue
        # The tracker sometimes reports thousands of formatted empty columns.
        rows = [list(row) for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                                         max_col=min(45, worksheet.max_column),
                                                         values_only=True)]
        header_index = find_header_row(rows)
        if header_index is None:
            continue
        column_map = {}
        for index, header in enumerate(rows[header_index]):
            canonical = canonical_column(header)
            if canonical and canonical not in column_map:
                column_map[canonical] = index
        if "Talent Email" not in column_map:
            inferred_email = infer_email_column(rows, header_index)
            if inferred_email is not None:
                column_map["Talent Email"] = inferred_email
        if "Applicant Name" not in column_map or "Talent Email" not in column_map:
            continue
        for row in rows[header_index + 1:]:
            name = row[column_map["Applicant Name"]] if column_map["Applicant Name"] < len(row) else None
            email = row[column_map["Talent Email"]] if column_map["Talent Email"] < len(row) else None
            if not normalize_name(name) and not normalize_email(email):
                continue
            record = {field: "" for field in FIELD_ALIASES}
            for field, index in column_map.items():
                record[field] = row[index] if index < len(row) and row[index] is not None else ""
            record["Tracker Sheet"] = worksheet.title
            record["_name"] = normalize_name(name)
            record["_email"] = normalize_email(email)
            records.append(record)
    return pd.DataFrame(records)


@st.cache_data(show_spinner=False)
def read_attendance(file_bytes):
    data = pd.read_excel(io.BytesIO(file_bytes), dtype=str).fillna("")
    columns = {clean_header(column): column for column in data.columns}
    name_col = columns.get("name")
    email_col = columns.get("email")
    number_col = columns.get("n")
    if not name_col or not email_col:
        raise ValueError("The attendance file must contain NAME and EMAIL columns.")
    result = pd.DataFrame({
        "N": data[number_col] if number_col else range(1, len(data) + 1),
        "NAME": data[name_col],
        "EMAIL": data[email_col],
    })
    result["_name"] = result["NAME"].map(normalize_name)
    result["_email"] = result["EMAIL"].map(normalize_email)
    return result


def is_complete(field, value):
    value = normalize_text(value)
    if field == "Fingerprint Screening":
        return value == "approved"
    if not value:
        return False
    incomplete_terms = (
        "need", "pending", "initiated", "requested", "scheduled", "will send",
        "info requested", "inconclusive", "positive", "fail", "retake", "void",
        "not cleared", "no results", "awaiting", "mailed",
    )
    if any(term in value for term in incomplete_terms):
        return False
    if field == "Education Requirement":
        return value not in {"n/a", "na", "none"}
    accepted = {
        "complete", "completed", "approved", "pass", "printed", "taken",
        "received", "recieved", "agent received", "yes", "n/a", "na",
    }
    return value in accepted


def best_match(attendee, tracker, threshold):
    email_matches = tracker[tracker["_email"].eq(attendee["_email"]) & tracker["_email"].ne("")]
    if not email_matches.empty:
        return email_matches.iloc[0], "Email match", 100
    name_matches = tracker[tracker["_name"].eq(attendee["_name"]) & tracker["_name"].ne("")]
    if not name_matches.empty:
        return name_matches.iloc[0], "Name match", 100
    if not attendee["_name"] or tracker.empty:
        return None, "Not found", 0
    scores = tracker["_name"].map(lambda name: round(SequenceMatcher(None, attendee["_name"], name).ratio() * 100))
    best_index = scores.idxmax()
    score = int(scores.loc[best_index])
    if score >= threshold:
        return tracker.loc[best_index], "Possible match – review", score
    return None, "Not found", score


def compare(attendance, tracker, threshold):
    output = []
    for _, attendee in attendance.iterrows():
        match, match_type, score = best_match(attendee, tracker, threshold)
        row = {"N": attendee["N"], "NHO Name": attendee["NAME"], "NHO Email": attendee["EMAIL"],
               "Match Result": match_type, "Match Score": score}
        if match is None:
            row.update({"Tracker Name": "", "Tracker Email": "", "Tracker Sheet": "",
                        "Current Onboarding Stage": "", "Missing Requirements": "Not found in tracker",
                        "Fingerprint Alert": "", "Notes": ""})
        else:
            missing = [field for field in REQUIREMENTS if not is_complete(field, match.get(field, ""))]
            fingerprint_value = match.get("Fingerprint Screening", "")
            row.update({"Tracker Name": match.get("Applicant Name", ""),
                        "Tracker Email": match.get("Talent Email", ""),
                        "Tracker Sheet": match.get("Tracker Sheet", ""),
                        "Current Onboarding Stage": match.get("Current Onboarding Stage", ""),
                        "Missing Requirements": ", ".join(missing) if missing else "None",
                        "Fingerprint Alert": "YES — ACTION REQUIRED" if not is_complete("Fingerprint Screening", fingerprint_value) else "",
                        "Notes": match.get("Notes", "")})
            for field in REQUIREMENTS:
                row[field] = match.get(field, "")
        output.append(row)
    return pd.DataFrame(output)


def make_excel(results):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        results.to_excel(writer, sheet_name="NHO Check", index=False)
        sheet = writer.book["NHO Check"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="4C9F38")
        alert_column = next(cell.column for cell in sheet[1] if cell.value == "Fingerprint Alert")
        red_fill = PatternFill(fill_type="solid", fgColor="7F1D1D")
        white_font = Font(color="FFFFFF", bold=True)
        for row in sheet.iter_rows(min_row=2):
            if row[alert_column - 1].value:
                for cell in row:
                    cell.fill = red_fill
                    cell.font = white_font
        for column in sheet.columns:
            width = min(50, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
    return buffer.getvalue()


def make_pdf(results):
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        rightMargin=0.35 * inch, leftMargin=0.35 * inch,
        topMargin=0.5 * inch, bottomMargin=0.45 * inch,
        title="NHO Attendance Comparison",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=18, leading=21, textColor=colors.HexColor("#253746"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"], fontSize=9, leading=11,
        textColor=colors.HexColor("#52616B"), spaceAfter=9,
    )
    cell_style = ParagraphStyle(
        "TableCell", parent=styles["BodyText"], fontSize=7, leading=8.5,
        textColor=colors.HexColor("#1F2933"),
    )
    header_style = ParagraphStyle(
        "TableHeader", parent=cell_style, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=TA_CENTER,
    )
    alert_cell_style = ParagraphStyle(
        "AlertTableCell", parent=cell_style, fontName="Helvetica-Bold",
        textColor=colors.white,
    )

    def paragraph(value, style=cell_style):
        safe = str(value if value is not None else "").replace("–", "-").replace("—", "-")
        safe = safe.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(safe, style)

    data = [[paragraph(value, header_style) for value in
             ["#", "NHO Name", "Match", "Tracker Sheet", "Stage", "Fingerprints", "Missing Processes"]]]
    for _, row in results.iterrows():
        row_style = alert_cell_style if row.get("Fingerprint Alert", "") == "YES — ACTION REQUIRED" else cell_style
        data.append([
            paragraph(row.get("N", ""), row_style), paragraph(row.get("NHO Name", ""), row_style),
            paragraph(row.get("Match Result", ""), row_style), paragraph(row.get("Tracker Sheet", ""), row_style),
            paragraph(row.get("Current Onboarding Stage", ""), row_style),
            paragraph(row.get("Fingerprint Screening", ""), row_style),
            paragraph(row.get("Missing Requirements", ""), row_style),
        ])

    table = Table(data, repeatRows=1,
                  colWidths=[0.32 * inch, 1.25 * inch, 0.85 * inch, 0.75 * inch,
                             1.35 * inch, 0.85 * inch, 3.83 * inch])
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4C9F38")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD2D9")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for index, (_, row) in enumerate(results.iterrows(), start=1):
        if row.get("Fingerprint Alert", "") == "YES — ACTION REQUIRED":
            commands.extend([
                ("BACKGROUND", (0, index), (-1, index), colors.HexColor("#7F1D1D")),
                ("TEXTCOLOR", (0, index), (-1, index), colors.white),
            ])
        elif index % 2 == 0:
            commands.append(("BACKGROUND", (0, index), (-1, index), colors.HexColor("#F4F7F2")))
    table.setStyle(TableStyle(commands))

    exact = int(results["Match Result"].isin(["Email match", "Name match"]).sum())
    alerts = int(results["Fingerprint Alert"].eq("YES — ACTION REQUIRED").sum())
    story = [
        Paragraph("NHO Attendance Comparison", title_style),
        Paragraph(f"Attendees: {len(results)} &nbsp;&nbsp; Exact matches: {exact} &nbsp;&nbsp; Fingerprint alerts: {alerts}", subtitle_style),
        table, Spacer(1, 4),
    ]

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#52616B"))
        canvas.drawRightString(10.65 * inch, 0.2 * inch, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    return buffer.getvalue()


st.title("NHO Attendance Checker")
st.caption("Compare an NHO attendance list with the Miami-Dade onboarding tracker.")

left, right = st.columns(2)
with left:
    attendance_file = st.file_uploader("1. Upload NHO attendance Excel", type=["xlsx"], key="attendance")
with right:
    tracker_file = st.file_uploader("2. Upload Miami-Dade Tracker Excel", type=["xlsx"], key="tracker")

threshold = st.slider("Minimum score for a possible name match", 70, 100, 88,
                      help="Possible matches are never treated as exact; review them before using the result.")

if attendance_file and tracker_file:
    try:
        with st.spinner("Reading and comparing files…"):
            attendance = read_attendance(attendance_file.getvalue())
            tracker = read_tracker(tracker_file.getvalue())
            results = compare(attendance, tracker, threshold)
        found = results["Match Result"].ne("Not found").sum()
        exact = results["Match Result"].isin(["Email match", "Name match"]).sum()
        review = results["Match Result"].eq("Possible match – review").sum()
        missing = results["Match Result"].eq("Not found").sum()
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Attendees", len(results)); c2.metric("Exact matches", int(exact))
        c3.metric("Review", int(review)); c4.metric("Not found", int(missing))
        st.success(f"Comparison complete: {found} of {len(results)} attendees matched the tracker.")
        filter_choice = st.segmented_control("Show", ["All", "Exact", "Review", "Not found"], default="All")
        shown = results
        if filter_choice == "Exact": shown = results[results["Match Result"].isin(["Email match", "Name match"])]
        elif filter_choice == "Review": shown = results[results["Match Result"].eq("Possible match – review")]
        elif filter_choice == "Not found": shown = results[results["Match Result"].eq("Not found")]
        def highlight_fingerprint_alert(row):
            if row.get("Fingerprint Alert", "") == "YES — ACTION REQUIRED":
                return ["background-color: #7f1d1d; color: #ffffff; font-weight: 600"] * len(row)
            return [""] * len(row)

        styled = shown.style.apply(highlight_fingerprint_alert, axis=1)
        st.dataframe(styled, use_container_width=True, hide_index=True,
                     column_config={"Match Score": st.column_config.ProgressColumn(min_value=0, max_value=100)})
        download_excel, download_pdf = st.columns(2)
        with download_excel:
            st.download_button("Download comparison as Excel", make_excel(results),
                               file_name="NHO_comparison.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               type="primary", use_container_width=True)
        with download_pdf:
            st.download_button("Download comparison as PDF", make_pdf(results),
                               file_name="NHO_comparison.pdf", mime="application/pdf",
                               use_container_width=True)
    except Exception as error:
        st.error(f"Could not compare the files: {error}")
else:
    st.info("Upload both Excel files to begin. Files are processed in memory and are not saved by the app.")
